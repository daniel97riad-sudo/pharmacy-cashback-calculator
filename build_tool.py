#!/usr/bin/env python3
"""
Build the pharmacy cashback calculator HTML tool from a merchandising Excel file.

Usage:
    python build_tool.py <path/to/merchandising.xlsx> [output.html]
    python build_tool.py <path/to/merchandising.xlsx> --dump-headers
    python build_tool.py <path/to/merchandising.xlsx> [output.html] --name-col E --q3-dist-col P ...

The Excel file is expected to have one row per pharmacy with columns
(header text match is fuzzy/case-insensitive, order doesn't matter) for:
  - pharmacy name              (e.g. "Ims_Customer_Came", "Customer Name", "Pharmacy Name")
  - customer code              (e.g. "Customer_Code")
  - MR / rep name               (e.g. "MR Name")
  - DM / district manager name (e.g. "DM Name")
  - region                     (e.g. "Region_Name")
  - territory                  (e.g. "Territory_Name")
  - pharmacy status            (e.g. "Pharmacy Status")
  - Q3 actual sales            (header containing "Q3" and "Actual")
  - Q3 distributor sales       (header containing "Q3" and "DIST")
  - Q4 actual sales            (header containing "Q4" and "Actual")
  - Q4 distributor sales       (header containing "Q4" and "DIST")

It scans the first 15 rows of the first sheet to find the header row
(whichever row matches the most expected labels), then reads every row
below it until it hits a run of empty pharmacy-name cells.

WHEN AUTO-DETECTION ISN'T ENOUGH: if a source file uses header wording this
script doesn't recognize, or has no clear headers at all, don't guess at the
mapping. Run with --dump-headers first to see exactly what's in the file
(row-by-row, column letter + text), ask the user which column is which, then
re-run passing the confirmed columns explicitly via --name-col, --code-col,
--mr-col, --dm-col, --region-col, --territory-col, --status-col,
--q3-actual-col, --q3-dist-col, --q4-actual-col, --q4-dist-col (each takes a
column letter like "E"), plus --header-row if the header row itself was
misdetected. Explicit overrides always win over auto-detection for that field.

Output: a single self-contained HTML file (Arabic RTL) with the pharmacy
data embedded, ready to open in any browser. No sales data leaves the
user's machine at build time -- everything is local file I/O.
"""
import sys
import json
import re
import argparse
import difflib
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:
    print("Missing dependency. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)

HEADER_PATTERNS = {
    "n":  [r"ims_customer_came", r"customer name", r"pharmacy name", r"customer_name"],
    "c":  [r"customer_code", r"^code$", r"customer code"],
    "mr": [r"mr name", r"\bmr\b.*name"],
    "dm": [r"dm name", r"\bdm\b.*name"],
    "rg": [r"region_name", r"^region$"],
    "tr": [r"territory_name", r"^territory$"],
    "st": [r"pharmacy status", r"^status$"],
}

# Maps CLI override flag names to the internal column keys used everywhere else.
OVERRIDE_KEYS = {
    "name_col": "n", "code_col": "c", "mr_col": "mr", "dm_col": "dm",
    "region_col": "rg", "territory_col": "tr", "status_col": "st",
    "q3_actual_col": "q3a", "q3_dist_col": "q3d",
    "q4_actual_col": "q4a", "q4_dist_col": "q4d",
}

def find_col(header_row, patterns):
    for cell_val, col_idx in header_row:
        if cell_val is None:
            continue
        text = str(cell_val).strip().lower()
        for pat in patterns:
            if re.search(pat, text):
                return col_idx
    return None

def find_quarter_cols(header_row, quarter_label):
    """Find the Actual and DIST sales columns for a given quarter (e.g. 'q3', 'q4')."""
    actual_col = dist_col = None
    for cell_val, col_idx in header_row:
        if cell_val is None:
            continue
        text = str(cell_val).strip().lower()
        if quarter_label in text and "actual" in text:
            actual_col = col_idx
        elif quarter_label in text and "dist" in text:
            dist_col = col_idx
    return actual_col, dist_col

def _read_all_rows(xlsx_path):
    """Load a workbook's first sheet in read-only mode and materialize every row as a
    tuple via iter_rows (sequential access). This is dramatically faster than repeated
    ws.cell(row, column) random-access lookups on large sheets -- that pattern alone
    was taking 45+ seconds on a ~2,750-row x 100+ column file. Returns (all_rows, max_col)
    where all_rows[r-1] is row r's tuple of values (1-indexed row numbers, 0-indexed tuple)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = [row for row in ws.iter_rows(values_only=True)]
    max_col = ws.max_column
    wb.close()
    return all_rows, max_col

def locate_header_row(all_rows, max_scan_rows=15):
    best_row, best_score = None, -1
    for r in range(1, min(max_scan_rows, len(all_rows)) + 1):
        row_vals = all_rows[r - 1]
        row_cells = list(enumerate(row_vals, start=1))
        row_cells = [(v, c) for c, v in row_cells]
        score = 0
        for patterns in HEADER_PATTERNS.values():
            if find_col(row_cells, patterns) is not None:
                score += 1
        q3a, q3d = find_quarter_cols(row_cells, "q3")
        q4a, q4d = find_quarter_cols(row_cells, "q4")
        score += sum(x is not None for x in [q3a, q3d, q4a, q4d])
        if score > best_score:
            best_score, best_row = score, r
    return best_row

def dump_headers(xlsx_path, max_scan_rows=15):
    """Print every non-empty header cell (row by row, column letter + text) so a
    human can see exactly what's in the file and say which column is which."""
    all_rows, max_col = _read_all_rows(xlsx_path)
    detected = locate_header_row(all_rows, max_scan_rows)

    print(f"(auto-detected header row: {detected})\n")
    for r in range(1, min(max_scan_rows, len(all_rows)) + 1):
        row_vals = all_rows[r - 1]
        nonblank = [(v, c) for c, v in enumerate(row_vals, start=1) if v is not None and str(v).strip() != ""]
        if not nonblank:
            continue
        marker = "  <-- auto-detected header row" if r == detected else ""
        print(f"Row {r}:{marker}")
        for v, c in nonblank:
            print(f"  {get_column_letter(c)}: {v!r}")
        print()

def extract_data(xlsx_path, overrides=None, forced_header_row=None):
    overrides = overrides or {}
    all_rows, max_col = _read_all_rows(xlsx_path)

    header_r = forced_header_row if forced_header_row else locate_header_row(all_rows)
    if header_r is None:
        raise RuntimeError(
            "Could not locate a header row in the first sheet. "
            "Run with --dump-headers to see the raw file contents, then re-run "
            "with --header-row and the --*-col overrides once you know which is which."
        )

    header_vals = all_rows[header_r - 1]
    row_cells = list(zip(header_vals, range(1, len(header_vals) + 1)))
    cols = {key: find_col(row_cells, pats) for key, pats in HEADER_PATTERNS.items()}
    q3a_col, q3d_col = find_quarter_cols(row_cells, "q3")
    q4a_col, q4d_col = find_quarter_cols(row_cells, "q4")
    auto_cols = {**cols, "q3a": q3a_col, "q3d": q3d_col, "q4a": q4a_col, "q4d": q4d_col}

    # Explicit column-letter overrides always win over auto-detection.
    for override_flag, key in OVERRIDE_KEYS.items():
        letter = overrides.get(override_flag)
        if letter:
            auto_cols[key] = column_index_from_string(letter.strip().upper())

    missing = [k for k in ("n", "q3d", "q4d") if auto_cols.get(k) is None]
    if missing:
        header_preview = "\n".join(
            f"  {get_column_letter(c)}: {v!r}" for v, c in row_cells if v is not None and str(v).strip() != ""
        )
        raise RuntimeError(
            f"Could not find required columns for: {missing}.\n"
            f"Header row {header_r} contains:\n{header_preview}\n\n"
            f"Don't guess -- ask the user which column above is which, then re-run with "
            f"the matching --name-col / --q3-dist-col / --q4-dist-col (etc.) flags using "
            f"the column letters shown."
        )

    def val(row_vals, col):
        if col is None or col > len(row_vals):
            return None
        return row_vals[col - 1]

    rows = []
    blank_streak = 0
    r = header_r + 1
    while r <= len(all_rows) and blank_streak < 25:
        row_vals = all_rows[r - 1]
        name = val(row_vals, auto_cols["n"])
        if name is None or str(name).strip() == "":
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        rows.append({
            "c": val(row_vals, auto_cols["c"]),
            "n": str(name).strip(),
            "mr": str(val(row_vals, auto_cols["mr"]) or "").strip(),
            "dm": str(val(row_vals, auto_cols["dm"]) or "").strip(),
            "rg": str(val(row_vals, auto_cols["rg"]) or "").strip(),
            "tr": str(val(row_vals, auto_cols["tr"]) or "").strip(),
            "st": str(val(row_vals, auto_cols["st"]) or "").strip(),
            "q3d": round(float(val(row_vals, auto_cols["q3d"]) or 0), 2),
            "q3a": round(float(val(row_vals, auto_cols["q3a"]) or 0), 2) if auto_cols["q3a"] else 0,
            "q4d": round(float(val(row_vals, auto_cols["q4d"]) or 0), 2),
            "q4a": round(float(val(row_vals, auto_cols["q4a"]) or 0), 2) if auto_cols["q4a"] else 0,
        })
        r += 1

    return rows

CHC_REQUIRED_COLS = [
    "customer_code", "ims_customer_name", "region_name", "territory_name",
    "mr_name", "dm_name", "ims_cust_type", "ph_flag", "merch_flag", "medical_flag",
]

def extract_chc_data(xlsx_path, ph_flag_only=True):
    """Extract pharmacies from a master CHC customer file -- a full multi-channel
    customer master (pharmacies, hospitals, universities, etc.) with one column per
    calendar month (V_YYYYMM = that month's distributor sales value) rather than
    pre-aggregated Q3/Q4 columns. Unlike extract_data (built for a merchandising-
    list shape with fixed Q3/Q4 Actual/DIST columns), this reads the header row
    literally by name -- this file's shape is fixed and well-known, not something
    that needs fuzzy column-guessing.

    ph_flag_only=True (the default) keeps only rows flagged as an actual pharmacy
    (ph_flag == TRUE) -- the source file also contains hospitals, universities,
    the army, etc. under the same customer master, which this tool has no reason
    to include.

    Q3 2025 / Q4 2025 are derived by summing that quarter's 3 monthly V_ columns
    (Jul+Aug+Sep, Oct+Nov+Dec) -- kept for the slab/cashback calculator baseline,
    unchanged from before. Separately, EVERY monthly V_ column the file has (Jan
    2024 through whatever the latest month present is -- currently Jun 2026) is
    kept per-pharmacy under "mo", keyed by yyyymm, purely for the sales-history
    chart: the chart renders whatever's in this dict, sorted chronologically, so
    a rep can see the full multi-year trend, not just the two quarters used for
    the cashback math. "ims_cust_type" is kept as "ct" (e.g. PHAR/CHPH/EGPH) for
    display. There's no equivalent of the old "actual sales (IMS)" figure or
    "pharmacy status" field in this file's shape, so those are left blank -- see
    build_tool.py's caller for how the template handles missing status.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {h: i for i, h in enumerate(header)}

    missing = [c for c in CHC_REQUIRED_COLS if c not in idx]
    if missing:
        raise RuntimeError(
            f"final_chc-style file is missing expected columns: {missing}.\n"
            f"Found headers: {list(idx.keys())}\n\n"
            f"Don't guess -- ask the user to confirm the file/column names before proceeding."
        )

    def val(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    def month_val(row, yyyymm):
        i = idx.get(f"V_{yyyymm}")
        if i is None or i >= len(row):
            return 0.0
        v = row[i]
        return float(v) if v else 0.0

    # Discover every monthly V_ column the file actually has, rather than hardcoding
    # a date range -- so future refreshes of this file (more months added) pick up
    # automatically without a code change. Sorted so the chart always plots them
    # left-to-right in chronological order regardless of column order in the sheet.
    all_month_cols = sorted(
        h[2:] for h in idx.keys() if isinstance(h, str) and h.startswith("V_") and h[2:].isdigit()
    )

    rows = []
    for row in rows_iter:
        ph = val(row, "ph_flag")
        is_ph = (ph is True) or (str(ph).strip().upper() == "TRUE")
        if ph_flag_only and not is_ph:
            continue
        code = val(row, "customer_code")
        name = val(row, "ims_customer_name")
        if code is None or name is None or str(name).strip() == "":
            continue

        q3d = round(sum(month_val(row, f"2025{m:02d}") for m in (7, 8, 9)), 2)
        q4d = round(sum(month_val(row, f"2025{m:02d}") for m in (10, 11, 12)), 2)
        mo = {yyyymm: round(month_val(row, yyyymm), 2) for yyyymm in all_month_cols}

        # merch_flag distinguishes pharmacies that were already part of the
        # merchandising program (the old ~2,751-pharmacy list this tool used to be
        # built from) from ones that are only showing up here because they're a
        # pharmacy in the CHC master (ph_flag=TRUE) but were never on the
        # merchandising list. Reuses the existing "st" exist/new badge field/UI --
        # "Exist" = PH & Merch (on both lists), "New" = PH only (this file only).
        mf_raw = val(row, "merch_flag")
        is_merch = (mf_raw is True) or (str(mf_raw).strip().upper() == "TRUE")

        rows.append({
            "c": str(code).strip(),
            "n": str(name).strip(),
            "mr": str(val(row, "mr_name") or "").strip(),
            "dm": str(val(row, "dm_name") or "").strip(),
            "rg": str(val(row, "region_name") or "").strip(),
            "tr": str(val(row, "territory_name") or "").strip(),
            "st": "Exist" if is_merch else "New",
            "ct": str(val(row, "ims_cust_type") or "").strip(),
            "q3d": q3d, "q3a": 0, "q4d": q4d, "q4a": 0,
            "mo": mo if mo else None,
        })
    wb.close()
    return rows

def strip_nonpharmacy_leaked_mr_names(xlsx_path, rows):
    """Some names appearing as mr_name on a handful of pharmacy rows are, in reality,
    reps from another channel entirely (medical/hospital, university, army, distributor
    stores, etc.) whose name leaked onto those rows by a data error in final_chc.xlsx --
    their real account book is overwhelmingly non-pharmacy (ph_flag=FALSE) accounts.
    Originally this only checked medical_flag=TRUE rows, but inspection showed several
    names are contaminated almost entirely from OTHER non-pharmacy channels with zero
    medical rows at all (e.g. "Ayman Henin": 4 pharmacy rows vs. 14 other-channel rows,
    0 medical) -- so the check now counts every non-pharmacy row, not just medical ones.
    Confirmed extreme case: "Yasmin Ehab Ibrahem Abed" sits on just 1 pharmacy row but
    207 non-pharmacy rows (8 medical + 199 other-channel). Meanwhile a legitimate
    pharmacy MR who also happens to touch a few non-pharmacy accounts (e.g. "Amr
    Hussein Kamel Ali Nasser": 191 pharmacy rows vs. a handful of others) is left
    untouched.

    Rule: for each mr_name, compare how many pharmacy rows (ph_flag=TRUE, already in
    `rows`) carry that name vs. how many non-pharmacy rows (ph_flag=FALSE, any channel)
    do. If the non-pharmacy count exceeds the pharmacy count, that name is not really a
    pharmacy MR -- blank out mr_name on those specific pharmacy rows (the pharmacy
    itself, its sales data, dm_name, etc. are left untouched; only the wrong mr_name
    field is cleared, so the tool shows "unassigned" instead of another channel's rep).
    In practice this correlates almost perfectly with the pharmacy row's dm_name also
    being blank (135 of 136 rows), but blank-dm_name is deliberately NOT used as an
    independent trigger -- it's also common on genuinely legitimate, high-volume
    pharmacy MRs (a data-entry gap, not evidence of a wrong name), so using it alone
    would incorrectly wipe real MRs off hundreds of rows.

    Returns a list of (name, pharmacy_row_count, nonpharmacy_row_count) for every name
    that got stripped, for the caller to report.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {h: i for i, h in enumerate(header)}

    def val(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    nonpharmacy_counts = {}
    for row in rows_iter:
        ph = val(row, "ph_flag")
        is_ph = (ph is True) or (str(ph).strip().upper() == "TRUE")
        if is_ph:
            continue
        mr = str(val(row, "mr_name") or "").strip()
        if mr:
            nonpharmacy_counts[mr] = nonpharmacy_counts.get(mr, 0) + 1
    wb.close()

    pharmacy_counts = {}
    for row in rows:
        mr = row.get("mr")
        if mr:
            pharmacy_counts[mr] = pharmacy_counts.get(mr, 0) + 1

    stripped = []
    for name, nonph_count in nonpharmacy_counts.items():
        ph_count = pharmacy_counts.get(name, 0)
        if ph_count and nonph_count > ph_count:
            stripped.append((name, ph_count, nonph_count))

    stripped_names = {name for name, _, _ in stripped}
    if stripped_names:
        for row in rows:
            if row.get("mr") in stripped_names:
                row["mr"] = ""

    return stripped

def load_mr_dm_override_map(xlsx_path, forced_header_row=None):
    """Load MR/DM name overrides from a merchandising-list-shaped file (e.g.
    merch8.xlsx), keyed by customer code. Reuses the same fuzzy header detection as
    extract_data (HEADER_PATTERNS' "c"/"mr"/"dm" entries), since this file has that
    shape, not the CHC master's shape.

    Why this exists: when the pharmacy list was rebuilt from final_chc.xlsx, its
    mr_name/dm_name columns turned out to disagree with merch8.xlsx for a large
    share of the ~2,466 pharmacies present in both files -- mostly spelling variants
    of the same person ("Mohamed Salah" vs "Mohammed Salah", "Ahmed Albakry" vs
    "Ahmed El- Bakry"), some blanked-out DM names, and a handful of what look like
    genuinely different people. Since the merchandising list is the source the team
    already trusts for rep assignment on those pharmacies, this lets a build
    override final_chc.xlsx's mr/dm with the merchandising list's version for any
    customer code found in both -- pharmacies only in the CHC master (not on the
    merch list) keep whatever final_chc.xlsx has, since there's no alternative.

    Returns {code: (mr_name, dm_name)}.
    """
    all_rows, max_col = _read_all_rows(xlsx_path)
    header_r = forced_header_row if forced_header_row else locate_header_row(all_rows)
    if header_r is None:
        raise RuntimeError(
            "Could not locate a header row in the MR/DM override file. "
            "Run with --dump-headers on that file to see its raw contents, then "
            "re-run with --mr-dm-override-header-row."
        )
    header_vals = all_rows[header_r - 1]
    row_cells = list(zip(header_vals, range(1, len(header_vals) + 1)))
    code_col = find_col(row_cells, HEADER_PATTERNS["c"])
    mr_col = find_col(row_cells, HEADER_PATTERNS["mr"])
    dm_col = find_col(row_cells, HEADER_PATTERNS["dm"])
    missing = [k for k, v in [("code", code_col), ("mr", mr_col), ("dm", dm_col)] if v is None]
    if missing:
        raise RuntimeError(
            f"MR/DM override file is missing expected columns: {missing}. "
            f"Run with --dump-headers on that file to see its raw contents."
        )

    def val(row_vals, col):
        if col is None or col > len(row_vals):
            return None
        return row_vals[col - 1]

    override = {}
    for r in range(header_r + 1, len(all_rows) + 1):
        row_vals = all_rows[r - 1]
        code = val(row_vals, code_col)
        if code is None or str(code).strip() == "":
            continue
        mr = val(row_vals, mr_col)
        dm = val(row_vals, dm_col)
        override[str(code).strip()] = (
            str(mr).strip() if mr else "",
            str(dm).strip() if dm else "",
        )
    return override

def load_trusted_rep_roster(xlsx_path, forced_header_row=None):
    """Load the full set of MR names and DM names that actually appear in a
    merchandising-list-shaped file (e.g. merch_list.xlsx, the "2.7K" list) -- the
    roster the team trusts as real, currently-active reps. Reuses the same fuzzy
    header detection as extract_data/load_mr_dm_override_map, since this file has
    that shape. Returns (trusted_mr_names, trusted_dm_names), both sets of strings.
    """
    all_rows, max_col = _read_all_rows(xlsx_path)
    header_r = forced_header_row if forced_header_row else locate_header_row(all_rows)
    if header_r is None:
        raise RuntimeError(
            "Could not locate a header row in the trusted roster file. "
            "Run with --dump-headers on that file to see its raw contents."
        )
    header_vals = all_rows[header_r - 1]
    row_cells = list(zip(header_vals, range(1, len(header_vals) + 1)))
    mr_col = find_col(row_cells, HEADER_PATTERNS["mr"])
    dm_col = find_col(row_cells, HEADER_PATTERNS["dm"])

    def val(row_vals, col):
        if col is None or col > len(row_vals):
            return None
        return row_vals[col - 1]

    trusted_mr, trusted_dm = set(), set()
    for r in range(header_r + 1, len(all_rows) + 1):
        row_vals = all_rows[r - 1]
        mr = val(row_vals, mr_col)
        dm = val(row_vals, dm_col)
        if mr and str(mr).strip():
            trusted_mr.add(str(mr).strip())
        if dm and str(dm).strip():
            trusted_dm.add(str(dm).strip())
    return trusted_mr, trusted_dm

def reconcile_dm_with_trusted_roster(rows, trusted_dm, similarity_cutoff=0.55):
    """DM names are a small, stable roster (5 people, each covering hundreds of
    pharmacies) -- so unlike MR names, a dm_name currently in use but not literally
    in the trusted roster is almost certainly the SAME person with a spelling
    difference between final_chc.xlsx and the merch list, not a different person.
    Confirmed cases: "Mohammed Salah" / "Mohamed Salah" (763 rows), "Nagat El- Sissi"
    / "Najat Alsisi" (761 rows), "Ahmed El- Bakry" / "Ahmed Albakry" (633 rows) --
    together nearly a third of the whole pharmacy list, which is why this is worth
    correcting rather than blanking.

    For each dm_name not in trusted_dm, finds the trusted name with the highest
    difflib.SequenceMatcher ratio; if that ratio clears similarity_cutoff, corrects
    dm_name to the trusted spelling. names with no close-enough match are left
    untouched (there's no safe assumption to make about them).

    Returns a list of (old_name, new_name, row_count) for every correction made.
    """
    counts = {}
    for row in rows:
        dm = row.get("dm")
        if dm and dm not in trusted_dm:
            counts[dm] = counts.get(dm, 0) + 1

    corrections_map = {}
    report = []
    for name, count in counts.items():
        best_name, best_ratio = None, 0.0
        for t in trusted_dm:
            ratio = difflib.SequenceMatcher(None, name.lower(), t.lower()).ratio()
            if ratio > best_ratio:
                best_ratio, best_name = ratio, t
        if best_ratio >= similarity_cutoff:
            corrections_map[name] = best_name
            report.append((name, best_name, count))

    if corrections_map:
        for row in rows:
            dm = row.get("dm")
            if dm in corrections_map:
                row["dm"] = corrections_map[dm]

    return report

def strip_mr_not_in_trusted_roster(rows, trusted_mr):
    """Per an explicit decision to treat any MR name not appearing anywhere in the
    trusted merch-list roster as unverified: blank mr_name on pharmacy rows whose
    current mr_name isn't literally one of the known/trusted MR names. Unlike the DM
    case, these were checked for spelling-variant matches against the trusted roster
    and found none with real token overlap -- they're distinct names, most likely
    legitimately new reps for pharmacies that were never on the old merch list (which
    only covered ~2,751 of the 7,015 pharmacies here). There's no data-driven evidence
    these are wrong the way the medical/non-pharmacy leak names were -- this is purely
    "not on the list we currently trust," applied per explicit instruction.

    Returns a list of (name, row_count) for every name that got stripped.
    """
    counts = {}
    for row in rows:
        mr = row.get("mr")
        if mr and mr not in trusted_mr:
            counts[mr] = counts.get(mr, 0) + 1

    stripped_names = set(counts.keys())
    if stripped_names:
        for row in rows:
            if row.get("mr") in stripped_names:
                row["mr"] = ""

    return sorted(counts.items(), key=lambda x: -x[1])

def canonicalize_rep_names(rows, fields=("mr", "dm")):
    """Merges near-duplicate rep names that differ only in case/spacing/hyphenation
    (e.g. "wafaa Gamal Fathy eldegwey" from final_chc.xlsx vs "Wafaa Gamal Fathy
    eldegwey" from merch8.xlsx -- same person, inconsistent casing between the two
    source files) into a single canonical spelling per field, so MR/DM filter
    dropdowns and the leaderboard don't show the same rep split across two entries.

    Within each normalized group, prefers whichever variant starts with an uppercase
    letter (proper name casing); if more than one does (or none do), prefers
    whichever exact spelling is used by more pharmacies, so the more common form
    wins. Returns {field: [(canonical, [all variants merged into it])]} listing only
    the groups that actually had more than one variant, for reporting.
    """
    def normalize(s):
        s = s.lower()
        s = re.sub(r"el[\s\-]+", "el", s)   # "El- " / "El " / "el-" -> "el"
        s = re.sub(r"[\s\-.]", "", s)        # strip remaining spaces/hyphens/dots
        return s.strip()

    report = {}
    for field in fields:
        counts = {}
        for row in rows:
            v = row.get(field)
            if v:
                counts[v] = counts.get(v, 0) + 1

        groups = {}
        for v in counts:
            groups.setdefault(normalize(v), []).append(v)

        canonical_map = {}
        merges = []
        for variants in groups.values():
            if len(variants) < 2:
                continue
            capitalized = [v for v in variants if v[:1].isupper()]
            pool = capitalized if capitalized else variants
            canonical = max(pool, key=lambda v: counts[v])
            for v in variants:
                if v != canonical:
                    canonical_map[v] = canonical
            merges.append((canonical, variants))

        if canonical_map:
            for row in rows:
                v = row.get(field)
                if v in canonical_map:
                    row[field] = canonical_map[v]
        report[field] = merges
    return report

def apply_mr_dm_override(rows, override_map):
    """Overwrite mr/dm fields in-place for any row whose customer code is present in
    override_map, leaving every other row untouched. Returns (matched, changed):
    matched = how many rows had a code found in the override file at all, changed =
    of those, how many actually had a different mr or dm value before the override
    (so the caller can report how much this really moved, not just how much overlap
    there was)."""
    matched = changed = 0
    for row in rows:
        ov = override_map.get(row["c"])
        if ov is None:
            continue
        matched += 1
        new_mr, new_dm = ov
        if row["mr"] != new_mr or row["dm"] != new_dm:
            changed += 1
        row["mr"] = new_mr
        row["dm"] = new_dm
    return matched, changed

SUPPLEMENT_HEADER_PATTERNS = {
    "code": [r"customer_code", r"^code$", r"customer code"],
    "value": [r"dist_value", r"distributor.*value", r"^value$"],
}

def load_supplement_map(xlsx_path, code_col_override=None, value_col_override=None, header_row=1):
    """Load a secondary, differently-shaped sales file (e.g. a weekly distributor
    extract) and sum its sales column per customer code. Returns {code: total_value}.

    This is intentionally separate from extract_data/HEADER_PATTERNS: a weekly/
    ad-hoc sales file usually won't share the merchandising list's column layout
    (no MR/DM, no Q3/Q4 slabs), so don't try to force it through the same auto-
    detection. If this can't confidently find a code + value column, it raises
    rather than guessing -- dump the file's headers and ask the user which
    columns are which, then pass --supplement-code-col / --supplement-value-col.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    row_cells = [(v, i + 1) for i, v in enumerate(header_cells)]

    code_col = column_index_from_string(code_col_override.strip().upper()) if code_col_override else find_col(row_cells, SUPPLEMENT_HEADER_PATTERNS["code"])
    value_col = column_index_from_string(value_col_override.strip().upper()) if value_col_override else find_col(row_cells, SUPPLEMENT_HEADER_PATTERNS["value"])

    if code_col is None or value_col is None:
        header_preview = "\n".join(f"  {get_column_letter(c)}: {v!r}" for v, c in row_cells if v is not None and str(v).strip() != "")
        raise RuntimeError(
            f"Could not confidently find a customer-code column and/or a sales-value column in "
            f"{xlsx_path}.\nHeader row {header_row} contains:\n{header_preview}\n\n"
            f"Don't guess -- ask the user which column is which, then re-run with "
            f"--supplement-code-col and --supplement-value-col using the column letters shown."
        )

    totals = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = row[code_col - 1]
        value = row[value_col - 1]
        if code is None or str(code).strip() == "":
            continue
        code = str(code).strip()
        totals[code] = totals.get(code, 0) + (float(value) if value else 0)
    wb.close()
    return totals

def merge_supplement(rows, supplement_map, field_key):
    """Attach a supplemental sales figure to each row by matching on customer code.
    Rows with no match in the supplement file get None (not 0), so the tool can
    show "no data for this period" instead of implying zero sales."""
    matched = 0
    for row in rows:
        code = str(row.get("c") or "").strip()
        if code in supplement_map:
            row[field_key] = round(supplement_map[code], 2)
            matched += 1
        else:
            row[field_key] = None
    return matched

def load_monthly_supplement_map(xlsx_path, code_col_override=None, header_row=1):
    """Load a file shaped with one customer-code column plus one column per month
    (header like 202601, 202602, ... -- YYYYMM as an int or string). Returns
    (monthly_map, months) where monthly_map = {code: {month_int: value}} and
    months is the sorted list of month ints found, so the caller/template can
    render a proper multi-point time series instead of a single lump figure."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    row_cells = [(v, i + 1) for i, v in enumerate(header_cells)]

    code_col = column_index_from_string(code_col_override.strip().upper()) if code_col_override else find_col(row_cells, SUPPLEMENT_HEADER_PATTERNS["code"])

    month_cols = {}
    for v, c in row_cells:
        if v is None:
            continue
        text = str(v).strip()
        m = re.match(r"^(20\d{2})(0[1-9]|1[0-2])$", text)
        if m:
            month_cols[c] = int(text)

    if code_col is None or not month_cols:
        header_preview = "\n".join(f"  {get_column_letter(c)}: {v!r}" for v, c in row_cells if v is not None and str(v).strip() != "")
        raise RuntimeError(
            f"Could not confidently find a customer-code column and/or YYYYMM month columns in "
            f"{xlsx_path}.\nHeader row {header_row} contains:\n{header_preview}\n\n"
            f"Don't guess -- ask the user which column is which, then re-run with "
            f"--s1-code-col and/or --s1-header-row."
        )

    monthly_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = row[code_col - 1]
        if code is None or str(code).strip() == "":
            continue
        code = str(code).strip()
        entry = monthly_map.setdefault(code, {})
        for c, month_int in month_cols.items():
            v = row[c - 1] if c - 1 < len(row) else None
            entry[month_int] = round(entry.get(month_int, 0) + (float(v) if v else 0), 2)
    wb.close()
    return monthly_map, sorted(month_cols.values())

def merge_monthly_supplement(rows, monthly_map, field_key="s1"):
    """Attach a {month_str: value} dict per row under field_key, None if the
    pharmacy's customer code has no match in the monthly file."""
    matched = 0
    for row in rows:
        code = str(row.get("c") or "").strip()
        if code in monthly_map:
            row[field_key] = {str(m): v for m, v in sorted(monthly_map[code].items())}
            matched += 1
        else:
            row[field_key] = None
    return matched

def build_html(rows, template_path, output_path, supplement_label=None, week1_label=None):
    template = Path(template_path).read_text(encoding="utf-8")
    data_json = json.dumps(rows, ensure_ascii=False, separators=(",", ":"))
    label_json = json.dumps(supplement_label, ensure_ascii=False) if supplement_label else "null"
    week1_label_json = json.dumps(week1_label, ensure_ascii=False) if week1_label else "null"
    html = (template
            .replace("__DATA__", data_json)
            .replace("__COUNT__", str(len(rows)))
            .replace("__SUPPLEMENT_LABEL__", label_json)
            .replace("__WEEK1_LABEL__", week1_label_json))
    Path(output_path).write_text(html, encoding="utf-8")

def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("xlsx_path", nargs="?",
                   help="Legacy merchandising-list Excel file (old fixed-Q3/Q4-column shape). "
                        "Omit this and use --chc-file instead when building from a master CHC "
                        "customer file (monthly V_YYYYMM columns, ph_flag, ims_cust_type).")
    p.add_argument("output_path", nargs="?", default="pharmacy_cashback_tool.html")
    p.add_argument("--chc-file",
                   help="Master CHC customer file (one row per customer across all channels, with "
                        "customer_code/ims_customer_name/region_name/territory_name/mr_name/dm_name/"
                        "ims_cust_type/ph_flag and monthly V_YYYYMM sales columns). When given, this "
                        "becomes the base pharmacy list (filtered to ph_flag=TRUE) instead of xlsx_path "
                        "-- Q3/Q4 2025 and the Jan-Jun 2026 monthly chart points are derived directly "
                        "from its V_ columns, so --s1-file is not needed alongside this.")
    p.add_argument("--chc-include-non-pharmacy", action="store_true",
                   help="Include rows where ph_flag is not TRUE too (hospitals, universities, etc. "
                        "from the same customer master). Off by default -- this tool is pharmacy-only.")
    p.add_argument("--mr-dm-override-file",
                   help="A merchandising-list-shaped file (e.g. merch8.xlsx) whose mr_name/dm_name "
                        "should win over --chc-file's, for any customer code found in both. Use this "
                        "when the CHC master's rep-assignment columns are known to be stale/inconsistent "
                        "relative to the merchandising list for pharmacies that were already on it.")
    p.add_argument("--mr-dm-override-header-row", type=int, default=None,
                   help="Force the header row for --mr-dm-override-file if auto-detection picks wrong.")
    p.add_argument("--dump-headers", action="store_true",
                   help="Print every header cell found (row by row, column letter + text) and exit. "
                        "Use this first when a file's structure is unfamiliar, so you can ask the user "
                        "which column is which instead of guessing.")
    p.add_argument("--header-row", type=int, default=None,
                   help="Force which row number is the actual header row, if auto-detection picks the wrong one.")
    p.add_argument("--name-col")
    p.add_argument("--code-col")
    p.add_argument("--mr-col")
    p.add_argument("--dm-col")
    p.add_argument("--region-col")
    p.add_argument("--territory-col")
    p.add_argument("--status-col")
    p.add_argument("--q3-actual-col")
    p.add_argument("--q3-dist-col")
    p.add_argument("--q4-actual-col")
    p.add_argument("--q4-dist-col")
    p.add_argument("--supplement-file",
                   help="Path to a secondary sales file (e.g. a weekly distributor extract) to merge in "
                        "by customer code, as an extra 'wk' figure per pharmacy alongside the Q3/Q4 data. "
                        "Does not affect growth %% or slab/cashback calculations -- purely informational.")
    p.add_argument("--supplement-label", default="بيانات إضافية",
                   help="Label shown in the tool for the supplemental figure (e.g. 'July 2026').")
    p.add_argument("--supplement-code-col")
    p.add_argument("--supplement-value-col")
    p.add_argument("--supplement-header-row", type=int, default=1)
    p.add_argument("--s1-file",
                   help="Path to a monthly-shaped file (customer code + one column per YYYYMM) to merge "
                        "in as each pharmacy's 's1' monthly sales series, shown as extra points on the "
                        "sales-history chart. Purely informational -- does not affect growth/slab/cashback.")
    p.add_argument("--s1-code-col")
    p.add_argument("--s1-header-row", type=int, default=1)
    p.add_argument("--week1-file",
                   help="Path to a week-1-of-month distributor extract (customer code + value column) to "
                        "merge in as each pharmacy's 'wk1' figure, used by the week-1 pacing card to "
                        "project a full-month (and full-quarter) run-rate. Re-run this script with an "
                        "updated --week1-file each time a new week-1 extract is dropped on Google Drive.")
    p.add_argument("--week1-label", default="الأسبوع الأول من الشهر",
                   help="Label shown in the tool for the week-1 pacing figure.")
    p.add_argument("--week1-code-col")
    p.add_argument("--week1-value-col")
    p.add_argument("--week1-header-row", type=int, default=1)
    return p.parse_args()

def main():
    args = parse_args()

    if args.dump_headers:
        dump_headers(args.xlsx_path)
        return

    template_path = Path(__file__).parent.parent / "assets" / "tool_template.html"

    if args.chc_file:
        rows = extract_chc_data(args.chc_file, ph_flag_only=not args.chc_include_non_pharmacy)
    else:
        if not args.xlsx_path:
            print("Provide either xlsx_path (legacy merchandising list) or --chc-file.", file=sys.stderr)
            sys.exit(1)
        overrides = {
            "name_col": args.name_col, "code_col": args.code_col,
            "mr_col": args.mr_col, "dm_col": args.dm_col,
            "region_col": args.region_col, "territory_col": args.territory_col,
            "status_col": args.status_col,
            "q3_actual_col": args.q3_actual_col, "q3_dist_col": args.q3_dist_col,
            "q4_actual_col": args.q4_actual_col, "q4_dist_col": args.q4_dist_col,
        }
        rows = extract_data(args.xlsx_path, overrides=overrides, forced_header_row=args.header_row)
    if not rows:
        print("No pharmacy rows extracted -- double check the source file structure.", file=sys.stderr)
        sys.exit(1)

    if args.mr_dm_override_file:
        override_map = load_mr_dm_override_map(
            args.mr_dm_override_file, forced_header_row=args.mr_dm_override_header_row,
        )
        matched, changed = apply_mr_dm_override(rows, override_map)
        print(
            f"Applied MR/DM override from {args.mr_dm_override_file}: "
            f"{matched}/{len(rows)} pharmacies matched by customer code, "
            f"{changed} actually had a different mr/dm value before the override."
        )

    # Always run this, regardless of source: even without an override file, a single
    # source file can itself carry inconsistent casing across rows. Cheap and only
    # ever merges exact-normalized matches, so it's safe to run unconditionally.
    rep_merges = canonicalize_rep_names(rows)
    for field, merges in rep_merges.items():
        for canonical, variants in merges:
            others = [v for v in variants if v != canonical]
            print(f"Merged duplicate {field} name(s) into \"{canonical}\": {others}")

    # Only meaningful for a CHC-shaped source (it needs ph_flag/medical_flag on the
    # same file) -- strip mr_name from pharmacy rows where the name is really a
    # medical-channel rep who leaked onto a handful of pharmacy rows by data error.
    if args.chc_file:
        stripped = strip_nonpharmacy_leaked_mr_names(args.chc_file, rows)
        for name, ph_count, nonph_count in stripped:
            print(
                f"Stripped non-pharmacy-leaked mr_name \"{name}\" from {ph_count} pharmacy "
                f"row(s) -- has {nonph_count} non-pharmacy row(s), so not a real pharmacy MR."
            )

    # Reconcile mr_name/dm_name against the trusted merch-list roster: DM spelling
    # variants get corrected (small, stable roster -- a non-matching name is almost
    # certainly the same person spelled differently), MR names not on the roster get
    # blanked (per explicit instruction -- no assumption made that they're wrong,
    # just not verified against the trusted list).
    if args.mr_dm_override_file:
        trusted_mr, trusted_dm = load_trusted_rep_roster(
            args.mr_dm_override_file, forced_header_row=args.mr_dm_override_header_row,
        )
        dm_corrections = reconcile_dm_with_trusted_roster(rows, trusted_dm)
        for old_name, new_name, count in dm_corrections:
            print(f"Corrected dm_name \"{old_name}\" -> \"{new_name}\" on {count} pharmacy row(s).")

        mr_stripped = strip_mr_not_in_trusted_roster(rows, trusted_mr)
        for name, count in mr_stripped:
            print(f"Stripped mr_name \"{name}\" from {count} pharmacy row(s) -- not on the trusted roster.")

    supplement_label = None
    if args.supplement_file:
        supplement_map = load_supplement_map(
            args.supplement_file,
            code_col_override=args.supplement_code_col,
            value_col_override=args.supplement_value_col,
            header_row=args.supplement_header_row,
        )
        matched = merge_supplement(rows, supplement_map, "wk")
        supplement_label = args.supplement_label
        print(f"Merged supplement file: {matched}/{len(rows)} pharmacies matched by customer code.")

    if args.s1_file:
        monthly_map, months = load_monthly_supplement_map(
            args.s1_file,
            code_col_override=args.s1_code_col,
            header_row=args.s1_header_row,
        )
        matched = merge_monthly_supplement(rows, monthly_map, "s1")
        print(f"Merged S1 monthly file: {matched}/{len(rows)} pharmacies matched by customer code. Months found: {months}")

    week1_label = None
    if args.week1_file:
        week1_map = load_supplement_map(
            args.week1_file,
            code_col_override=args.week1_code_col,
            value_col_override=args.week1_value_col,
            header_row=args.week1_header_row,
        )
        matched = merge_supplement(rows, week1_map, "wk1")
        week1_label = args.week1_label
        print(f"Merged week-1 file: {matched}/{len(rows)} pharmacies matched by customer code.")

    build_html(rows, template_path, args.output_path, supplement_label=supplement_label, week1_label=week1_label)
    print(f"Extracted {len(rows)} pharmacies. Wrote {args.output_path}")

if __name__ == "__main__":
    main()
